from fastapi import APIRouter, HTTPException, Query, Form, UploadFile, File
from typing import Optional, Any
import csv
import io
from app.services import execution_service
from app.services import ai_service

router = APIRouter(prefix="/test-execution", tags=["Test Execution"])


def _parse_csv(file_bytes: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    reader = csv.DictReader(io.StringIO(file_bytes.decode("utf-8", errors="replace")))
    rows = list(reader)
    cols = list(reader.fieldnames or [])
    return cols, rows[:50]


def _parse_excel(file_bytes: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise HTTPException(
            status_code=422,
            detail="openpyxl is not installed on the server. Use CSV format instead.",
        ) from exc
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    headers = [str(c.value or f"col{i}") for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))]
    rows_out: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, max_row=52, values_only=True):
        rows_out.append(dict(zip(headers, row)))
    return headers, rows_out


@router.post("/run", status_code=201)
async def run_tests(
    data_source: str = Form(default="auto"),
    requirement_id: Optional[str] = Form(default=None),
    columns: Optional[str] = Form(default=None),
    preview_rows: Optional[str] = Form(default=None),
    file: Optional[UploadFile] = File(default=None),
):
    """
    Run all (or requirement-scoped) test cases.
    data_source: 'auto' | 'file' | 'columns'
    """
    try:
        data_context: dict[str, Any] | None = None

        if data_source == "file" and file and file.filename:
            file_bytes = await file.read()
            fname = file.filename.lower()
            if fname.endswith(".csv"):
                cols, rows = _parse_csv(file_bytes)
            elif fname.endswith((".xlsx", ".xls")):
                cols, rows = _parse_excel(file_bytes)
            else:
                raise HTTPException(status_code=422, detail="Unsupported file type. Upload a .csv or .xlsx file.")
            data_context = {
                "type": "file",
                "filename": file.filename,
                "columns": cols,
                "row_count": len(rows),
                "sample": rows[:5],
            }

        elif data_source == "columns" and columns:
            col_list = [c.strip() for c in columns.split(",") if c.strip()]
            if not col_list:
                raise HTTPException(status_code=422, detail="No column names provided.")
            # Ask Gemini to generate rows for those columns
            rows = await ai_service.generate_column_based_data(col_list, count=20)
            data_context = {
                "type": "columns",
                "columns": col_list,
                "row_count": len(rows),
                "sample": rows[:5],
            }

        elif data_source == "auto" and preview_rows:
            import json
            try:
                parsed_rows: list[dict[str, Any]] = json.loads(preview_rows)
            except json.JSONDecodeError as exc:
                raise HTTPException(status_code=422, detail="preview_rows is not valid JSON") from exc
            col_list = list(parsed_rows[0].keys()) if parsed_rows else []
            data_context = {
                "type": "auto",
                "columns": col_list,
                "row_count": len(parsed_rows),
                "sample": parsed_rows[:5],
            }

        summary = await execution_service.start_run(
            requirement_id=requirement_id,
            data_context=data_context,
        )
        # Return run_id + total immediately; results stream in via GET /results
        result = summary.model_dump()
        result["data_source"] = data_source
        result["data_columns"] = data_context["columns"] if data_context else []
        result["data_row_count"] = data_context["row_count"] if data_context else None
        return result

    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@router.get("/results")
async def get_results(
    run_id: Optional[str] = Query(default=None),
    limit: int = Query(default=100, le=500),
):
    return await execution_service.get_results(run_id=run_id, limit=limit)


@router.get("/runs/{run_id}")
async def get_run_summary(run_id: str):
    summary = await execution_service.get_run_summary(run_id)
    if not summary:
        raise HTTPException(status_code=404, detail=f"Run {run_id} not found")
    return summary
